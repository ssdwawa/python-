from openpyxl import load_workbook
from openpyxl import Workbook
wb3 = Workbook()
wb1 = load_workbook("1.xlsx")
wb2 = load_workbook("2.xlsx")
# sheet1 = wb1["苹果表2"]
# sheet1['A'][0].value='产品数量'
# wb1.save("1.xlsx")
sheet1 = wb1["苹果表2"]
sheet2 = wb2["橘子表2"]
dic1 = {}
dic2 = {}
temp_list = []
for i in range(2,sheet1.max_row+1):
    w1 = sheet1.cell(row=i, column=1).value
    w2 = sheet1.cell(row=i, column=2).value
    temp_list.append([w1,w2])
print(temp_list)
ws=wb3.active
for i in range(1,len(temp_list)):
    ws.cell(row=i, column=1).value=temp_list[i][0]
    ws.cell(row=i, column=2).value = temp_list[i][1]
wb3.save("sample.xlsx")