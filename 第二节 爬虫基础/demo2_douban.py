import urllib.request as req
import  re
from openpyxl import load_workbook
from openpyxl import Workbook
pat='<div class=\"name\">(.*?)</div>'
data=req.urlopen('https://read.douban.com/provider/all').read()
arr=re.compile(pat).findall(str(data,"utf-8"))
# wb = Workbook()
# ws=wb.active
# ws.cell(row=1, column=1).value='出版社'
# for i in range(2,len(arr)):
#     ws.cell(row=i, column=1).value = arr[i]
# wb.save("出版社.xlsx")
fh=open('file.txt','w')
for i in arr:
    fh.writelines(i + "\n")
fh.close()
